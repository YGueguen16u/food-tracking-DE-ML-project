import os
import requests
from googleapiclient.discovery import build
from openpyxl import load_workbook
import time

def download_images(url, save_path):
    try:
        response = requests.get(url, timeout=60)  # Ajouter un délai d'attente de 10 secondes
        response.raise_for_status()
    except RequestException as e:
        print(f"Erreur lors du téléchargement de l'image : {e}")
        return
    with open(save_path, 'wb') as file:
        for chunk in response.iter_content(chunk_size=8192):
            file.write(chunk)

def fetch_food_images(query, api_key, cx, num_images=10):
    service = build("customsearch", "v1", developerKey=api_key)

    image_urls = []
    for i in range(0, num_images, 10):
        start = i + 1
        response = service.cse().list(
            q=query,
            cx=cx,
            searchType='image',
            start=start,
            num=10
        ).execute()

        if 'items' not in response:
            print(f"Aucun résultat trouvé pour {query}")  # Ajout d'un message si aucun résultat n'est trouvé
            break

        for item in response['items']:
            image_urls.append(item['link'])

    return image_urls

def read_aliments_from_excel(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    aliments = []

    # Lire d'abord les fruits (colonne 1)
    for row in range(8, sheet.max_row + 1):
        fruit = sheet.cell(row=row, column=1).value
        if fruit:
            aliments.append((fruit, "fruit"))

    # Ensuite, lire les légumes (colonne 2)
    for row in range(5, sheet.max_row + 1):
        legume = sheet.cell(row=row, column=2).value
        if legume:
            aliments.append((legume, "légume"))

    return aliments

def main():
    # Remplacez ces valeurs par vos propres informations d'API
    api_key = "AIzaSyDPM8ZNAZAK4WvjlDa6jhRtowT_4q3auLg"
    cx = "56cf76f6353414099"

    num_images = 100
    save_directory = "food_imagess"

    os.makedirs(save_directory, exist_ok=True)

    excel_file = "aliment.xlsx"
    aliments = read_aliments_from_excel(excel_file)

    for aliment, keyword in aliments:
        print(f"Téléchargement des images pour {aliment}...")
        query = aliment + " " + keyword

        # Créez un sous-dossier pour chaque aliment
        aliment_save_directory = os.path.join(save_directory, aliment)
        os.makedirs(aliment_save_directory, exist_ok=True)

        image_urls = fetch_food_images(query, api_key, cx, num_images=num_images)

        for i, url in enumerate(image_urls):
            save_path = os.path.join(aliment_save_directory, f"{aliment}_image_{i+1}.jpg")
            try:
                download_images(url, save_path)
                print(f"Image {i+1} de {aliment} téléchargée avec succès.")
                time.sleep(1)  # Ajouter un délai d'une seconde entre les téléchargements
            except Exception as e:
                print(f"Erreur lors du téléchargement de l'image {i+1} de {aliment}: {e}")

if __name__ == "__main__":
    main()
